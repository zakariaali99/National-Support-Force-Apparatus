import io

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase

from openpyxl import load_workbook
from pypdf import PdfReader
from rest_framework.test import APIClient

from apps.core.models import Role, User
from apps.members.models import Member, MemberDocument, MemberNote
from apps.organization.models import DocumentType, Faction, Rank


class ReportsTestCase(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.rank = Rank.objects.create(code="rep-rank", name_ar="رتبة")
        self.faction = Faction.objects.create(code="rep-faction", name_ar="فصيل")
        self.other_faction = Faction.objects.create(code="rep-faction-2", name_ar="فصيل آخر")

        self.member = Member.objects.create(
            first_name="خالد",
            second_name="محمود",
            last_name="السنوسي",
            force_number="R-1",
            national_number="723456789012",
            rank=self.rank,
            faction=self.faction,
        )
        MemberNote.objects.create(member=self.member, body="ملاحظة للطباعة")

        printer_role = Role.objects.create(
            name="printer",
            name_ar="طابع",
            permissions=["member.view", "member.print", "member.export"],
            scope="all",
        )
        self.printer = User.objects.create_user(username="printer", password="x")
        self.printer.roles.add(printer_role)

        viewer_role = Role.objects.create(
            name="print-viewer", name_ar="مطّلع", permissions=["member.view"], scope="all"
        )
        self.viewer = User.objects.create_user(username="print-viewer", password="x")
        self.viewer.roles.add(viewer_role)

        scoped_role = Role.objects.create(
            name="scoped-printer",
            name_ar="طابع محدود",
            permissions=["member.view", "member.print", "member.export"],
            scope="own_faction",
        )
        self.scoped_printer = User.objects.create_user(username="scoped-printer", password="x")
        self.scoped_printer.roles.add(scoped_role)
        self.scoped_printer.factions.add(self.other_faction)


class SectionsListTests(ReportsTestCase):
    def test_any_authenticated_user_can_list_sections(self):
        self.client.force_authenticate(self.viewer)

        response = self.client.get("/api/reports/sections/")

        self.assertEqual(response.status_code, 200)
        keys = {row["key"] for row in response.data}
        self.assertIn("profile", keys)


class MemberPrintTests(ReportsTestCase):
    def test_requires_member_print_permission(self):
        self.client.force_authenticate(self.viewer)

        response = self.client.get(f"/api/members/{self.member.id}/print/?sections=profile")

        self.assertEqual(response.status_code, 403)

    def test_faction_scoped_user_cannot_print_other_faction_member(self):
        self.client.force_authenticate(self.scoped_printer)

        response = self.client.get(f"/api/members/{self.member.id}/print/?sections=profile")

        self.assertEqual(response.status_code, 403)

    def test_unknown_section_key_is_400(self):
        self.client.force_authenticate(self.printer)

        response = self.client.get(f"/api/members/{self.member.id}/print/?sections=profile,not-a-section")

        self.assertEqual(response.status_code, 400)

    def test_single_section_produces_one_page_pdf(self):
        self.client.force_authenticate(self.printer)

        response = self.client.get(f"/api/members/{self.member.id}/print/?sections=profile")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        reader = PdfReader(io.BytesIO(response.content))
        self.assertEqual(len(reader.pages), 1)

    def test_multiple_sections_produce_one_page_each(self):
        self.client.force_authenticate(self.printer)

        response = self.client.get(f"/api/members/{self.member.id}/print/?sections=profile,notes,tasks")

        self.assertEqual(response.status_code, 200)
        reader = PdfReader(io.BytesIO(response.content))
        self.assertEqual(len(reader.pages), 3)

    def test_download_flag_sets_attachment_disposition(self):
        self.client.force_authenticate(self.printer)

        response = self.client.get(f"/api/members/{self.member.id}/print/?sections=profile&download=1")

        self.assertIn("attachment", response["Content-Disposition"])

    def test_preview_flag_returns_html_instead_of_pdf(self):
        self.client.force_authenticate(self.printer)

        response = self.client.get(f"/api/members/{self.member.id}/print/?sections=profile&preview=1")

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"<html", response.content)

    def test_national_number_appears_in_rendered_pdf_text(self):
        self.client.force_authenticate(self.printer)

        response = self.client.get(f"/api/members/{self.member.id}/print/?sections=profile")

        reader = PdfReader(io.BytesIO(response.content))
        text = reader.pages[0].extract_text()
        self.assertIn("723456789012", text)

    def test_unknown_document_id_is_400(self):
        self.client.force_authenticate(self.printer)

        response = self.client.get(
            f"/api/members/{self.member.id}/print/?sections=profile&documents=999999"
        )

        self.assertEqual(response.status_code, 400)

    def test_image_document_included_adds_a_page(self):
        from PIL import Image

        buf = io.BytesIO()
        Image.new("RGB", (20, 20), color="white").save(buf, format="PNG")
        doc_type = DocumentType.objects.create(code="rep-doc", name_ar="مستند")
        document = MemberDocument.objects.create(
            member=self.member,
            document_type=doc_type,
            file=SimpleUploadedFile("scan.png", buf.getvalue(), content_type="image/png"),
            original_name="scan.png",
            content_type="image/png",
            file_size=len(buf.getvalue()),
            sha256="x" * 64,
        )
        self.client.force_authenticate(self.printer)

        response = self.client.get(
            f"/api/members/{self.member.id}/print/?sections=profile&documents={document.id}"
        )

        self.assertEqual(response.status_code, 200)
        reader = PdfReader(io.BytesIO(response.content))
        self.assertEqual(len(reader.pages), 2)


class MemberIdCardsTests(ReportsTestCase):
    def test_batch_id_cards_one_page_per_member(self):
        second = Member.objects.create(
            first_name="أمين",
            second_name="سالم",
            last_name="الفيتوري",
            force_number="R-2",
            national_number="723456789013",
            rank=self.rank,
            faction=self.faction,
        )
        self.client.force_authenticate(self.printer)

        response = self.client.get(f"/api/members/id-cards/?ids={self.member.id},{second.id}")

        self.assertEqual(response.status_code, 200)
        reader = PdfReader(io.BytesIO(response.content))
        self.assertEqual(len(reader.pages), 2)

    def test_faction_scoped_user_cannot_include_inaccessible_member(self):
        self.client.force_authenticate(self.scoped_printer)

        response = self.client.get(f"/api/members/id-cards/?ids={self.member.id}")

        self.assertEqual(response.status_code, 400)

    def test_missing_ids_param_is_400(self):
        self.client.force_authenticate(self.printer)

        response = self.client.get("/api/members/id-cards/")

        self.assertEqual(response.status_code, 400)


class MemberExportTests(ReportsTestCase):
    def test_export_requires_permission(self):
        self.client.force_authenticate(self.viewer)

        response = self.client.get("/api/members/export/")

        self.assertEqual(response.status_code, 403)

    def test_export_returns_xlsx_with_expected_row(self):
        self.client.force_authenticate(self.printer)

        response = self.client.get("/api/members/export/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[0][1], "الرقم الحربي")
        force_numbers = {row[1] for row in rows[1:]}
        self.assertIn("R-1", force_numbers)

    def test_export_respects_faction_filter(self):
        Member.objects.create(
            first_name="سليم",
            second_name="فتحي",
            last_name="الأصفر",
            force_number="R-3",
            national_number="723456789014",
            rank=self.rank,
            faction=self.other_faction,
        )
        self.client.force_authenticate(self.printer)

        response = self.client.get(f"/api/members/export/?faction={self.faction.id}")

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        force_numbers = {row[1] for row in ws.iter_rows(min_row=2, values_only=True)}
        self.assertEqual(force_numbers, {"R-1"})

    def test_export_scoped_to_user_factions(self):
        self.client.force_authenticate(self.scoped_printer)

        response = self.client.get("/api/members/export/")

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(rows, [])
